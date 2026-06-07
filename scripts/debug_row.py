"""Debug single row mismatch."""
import openpyxl
from pathlib import Path

base = Path(__file__).resolve().parents[1]
calc = list(base.glob("*28*.xlsx"))[0]
bi = list(base.glob("*BI*.xlsx"))[0]
out = base / "output" / "apuracao.xlsx"

forn, nota, prod = "211617", "000004371", "20904439009400"

for label, path, sheet_fn in [
    ("BI", bi, lambda wb: wb[wb.sheetnames[0]]),
    ("REF", calc, lambda wb: [wb[s] for s in wb.sheetnames if s.strip().startswith("DIFAL")][0]),
    ("OUT", out, lambda wb: [wb[s] for s in wb.sheetnames if s.strip().startswith("DIFAL")][0]),
]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = sheet_fn(wb)
    if label == "BI":
        hdr = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        idx = {str(h): i for i, h in enumerate(hdr) if h}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if str(row[idx["Cod Fornecedor"]]) == forn and str(row[idx["Nota Fiscal"]]).zfill(9) == nota and str(row[idx["Cód Produto"]]) == prod:
                print(f"\n{label}:")
                for k in ["Valor Contábil", "Carga Efetiva DIFAL", "D1_ICMSCOM", "NCM", "Cód Fiscal", "Desc. Grupo Produtos"]:
                    if k in idx:
                        print(f"  {k}: {row[idx[k]]}")
                break
    else:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == forn and str(row[3]) == nota and str(row[5]) == prod:
                print(f"\n{label}: VC={row[17]} ICMSCOM={row[22]} NOVO={row[23]} AJUSTE={row[24]} CONTA={row[4]}")
                break
    wb.close()
