"""Compare reference columns X, Y, Z, AA for NCM rows."""
import openpyxl
from pathlib import Path

calc = list(Path(__file__).resolve().parents[1].glob("*28*.xlsx"))[0]
wb_f = openpyxl.load_workbook(calc, read_only=False, data_only=False)
wb_d = openpyxl.load_workbook(calc, read_only=True, data_only=True)
ws_f = wb_f[[s for s in wb_f.sheetnames if "DIFAL" in s][0]]
ws_d = wb_d[[s for s in wb_d.sheetnames if "DIFAL" in s][0]]

for r in range(2, 350):
    zf = ws_f[f"Z{r}"].value
    if not (isinstance(zf, str) and zf.startswith("=")):
        continue
    row = list(ws_d.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    ncm = row[7]
    print(
        f"row{r} ncm={ncm} W={row[22]} X={row[23]} Y={row[24]} "
        f"Zval={row[25] if len(row)>25 else None} "
        f"AAval={row[26] if len(row)>26 else None} Zf={zf} "
        f"AAf={ws_f[f'AA{r}'].value}"
    )

wb_f.close()
wb_d.close()
