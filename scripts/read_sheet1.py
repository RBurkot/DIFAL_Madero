"""Inspect Sheet1 lookup table in reference."""
import openpyxl
from pathlib import Path

base = Path(__file__).resolve().parents[1]
calc = list(base.glob("*28*.xlsx"))[0]
wb = openpyxl.load_workbook(calc, read_only=True, data_only=True)

for name in wb.sheetnames:
    if "sheet" in name.lower() or name == "Planilha1":
        ws = wb[name]
        print(f"\n=== {name} ===")
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
            print(f"row{i+1}: {row[:25]}")
        # column 18 (index 17) for VLOOKUP col 18
        hdr = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        print(f"Col 18 header: {hdr[17] if len(hdr) > 17 else 'N/A'}")
        print(f"Col 3 (C) header: {hdr[2] if len(hdr) > 2 else 'N/A'}")

wb.close()
