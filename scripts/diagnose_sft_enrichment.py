"""Diagnóstico SFT: onde o enriquecimento (chave, nome, datas) falha.

Uso:
  python scripts/diagnose_sft_enrichment.py <planilha_ref_ou_apuracao.xlsx> [planilha_sft_opcional]

Se a planilha tiver aba DIFAL + SFT, um argumento basta.
Se SFT estiver só na referência *28*.xlsx, passe como 2º argumento.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from difal_importacao.reader import (  # noqa: E402
    build_chave_rastreio,
    lookup_fornecedor_by_chave,
    read_auxiliar_sft,
    read_difal,
)

SFT_CHAVE_COL = column_index_from_string("A") - 1
SFT_NOME_COL = column_index_from_string("I") - 1


def _sft_layout(path: Path) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((n for n in wb.sheetnames if n.strip().upper() == "SFT"), None)
    if not sheet:
        print(f"  Aba SFT não encontrada em {path.name}")
        wb.close()
        return
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        cells = list(row or [])
        chave = cells[SFT_CHAVE_COL] if len(cells) > SFT_CHAVE_COL else None
        nome = cells[SFT_NOME_COL] if len(cells) > SFT_NOME_COL else None
        preview = [str(c)[:20] if c is not None else "" for c in cells[:15]]
        print(f"  Linha {i}: A={chave!r} I={nome!r} | {preview}")
    wb.close()


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1

    entrada = Path(sys.argv[1])
    sft_path = Path(sys.argv[2]) if len(sys.argv) > 2 else entrada

    if not entrada.exists():
        print(f"Arquivo não encontrado: {entrada}")
        return 1

    print("=== 1. Layout SFT (primeiras 5 linhas, col A=Chave, col I=Nm Forn) ===")
    _sft_layout(sft_path)

    print("\n=== 2. Índice SFT ===")
    lookups = read_auxiliar_sft(sft_path, "SFT ")
    print(f"  Chaves indexadas: {len(lookups)}")
    if lookups:
        sample = next(iter(lookups.values()))
        print(f"  Exemplo: chave={sample.chave_rastreio!r} nome={sample.nome_fornecedor!r} "
              f"emissao={sample.data_emissao} entrada={sample.data_entrada}")

    print("\n=== 3. Linhas DIFAL elegíveis (ajuste != 0) ===")
    periodo, linhas = read_difal(entrada)
    elegiveis = [l for l in linhas if l.ajuste != 0]
    print(f"  Período: {periodo.label} | Total DIFAL: {len(linhas)} | Com ajuste: {len(elegiveis)}")

    hit = nome_ok = datas_ok = 0
    miss_samples: list[str] = []

    for linha in elegiveis:
        chave = build_chave_rastreio(linha.nota_fiscal, linha.fornecedor_cod, linha.cod_produto)
        lookup = lookup_fornecedor_by_chave(lookups, chave)
        if lookup:
            hit += 1
            if lookup.nome_fornecedor:
                nome_ok += 1
            if lookup.data_emissao and lookup.data_entrada:
                datas_ok += 1
        elif len(miss_samples) < 5:
            miss_samples.append(
                f"  nota={linha.nota_fiscal} forn={linha.fornecedor_cod!r} prod={linha.cod_produto} "
                f"→ chave={chave!r}"
            )

    n = len(elegiveis) or 1
    print(f"\n=== 4. Match DIFAL → SFT (chave col K = SFT col A) ===")
    print(f"  Chave encontrada:     {hit}/{len(elegiveis)} ({100 * hit / n:.1f}%)")
    print(f"  Nome preenchido (col I): {nome_ok}/{len(elegiveis)} ({100 * nome_ok / n:.1f}%)")
    print(f"  Datas L+M preenchidas: {datas_ok}/{len(elegiveis)} ({100 * datas_ok / n:.1f}%)")

    if miss_samples:
        print("\n  Amostras SEM match na SFT:")
        print("\n".join(miss_samples))
        print("\n  Dica: compare chave gerada com col A da SFT (formato fornecedor/produto).")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
