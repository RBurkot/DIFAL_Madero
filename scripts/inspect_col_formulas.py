"""Inspect Excel column letters and formulas in reference DIFAL."""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

calc = list(Path(__file__).resolve().parents[1].glob("*28*.xlsx"))[0]
wb = openpyxl.load_workbook(calc, read_only=False, data_only=False)
ws = wb[[s for s in wb.sheetnames if "DIFAL" in s][0]]
hdr = [c.value for c in ws[1]]
for i, h in enumerate(hdr, 1):
    print(get_column_letter(i), i, h)
print("--- formulas row 2 ---")
for col in range(18, 27):
    L = get_column_letter(col)
    print(L, ws[f"{L}2"].value)
print("Z1", ws["Z1"].value, "AA1", ws["AA1"].value)
for r in [2, 4, 18, 49, 50, 168]:
    print(
        f"row{r}: Z={ws[f'Z{r}'].value} AA={ws[f'AA{r}'].value} "
        f"R={ws[f'R{r}'].value} V={ws[f'V{r}'].value} W={ws[f'W{r}'].value} X={ws[f'X{r}'].value}"
    )

wb_data = openpyxl.load_workbook(calc, read_only=True, data_only=True)
ws2 = wb_data[[s for s in wb_data.sheetnames if "DIFAL" in s][0]]
for r in [4, 49, 50, 168]:
    row = list(ws2.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    z_val = row[25] if len(row) > 25 else None
    aa_val = row[26] if len(row) > 26 else None
    print(f"row{r} values: W={row[22]} X={row[23]} Y={row[24]} Z={z_val} AA={aa_val}")
wb.close()
wb_data.close()

# Scan all Z formulas
wb = openpyxl.load_workbook(calc, read_only=False, data_only=False)
ws = wb[[s for s in wb.sheetnames if "DIFAL" in s][0]]
z_formulas = set()
for r in range(2, 400):
    z = ws[f"Z{r}"].value
    if isinstance(z, str) and z.startswith("="):
        z_formulas.add(z)
print("Unique Z formulas:", z_formulas)
aa_formulas = set()
for r in range(2, 400):
    aa = ws[f"AA{r}"].value
    if isinstance(aa, str) and aa.startswith("="):
        aa_formulas.add(aa)
print("Unique AA formulas:", aa_formulas)
# Compare Z calc vs X for NCM rows
wsd = openpyxl.load_workbook(calc, read_only=True, data_only=True)[[s for s in openpyxl.load_workbook(calc, read_only=True).sheetnames if "DIFAL" in s][0]]
wb.close()
