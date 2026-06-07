"""Map grupo produto -> conta from reference + BI."""
import openpyxl
from pathlib import Path
from collections import defaultdict

base = Path(__file__).resolve().parents[1]
calc = list(base.glob("*28*.xlsx"))[0]
bi = list(base.glob("*BI*.xlsx"))[0]

wb_c = openpyxl.load_workbook(calc, read_only=True, data_only=True)
wb_b = openpyxl.load_workbook(bi, read_only=True, data_only=True)

difal_sheet = [s for s in wb_c.sheetnames if s.strip().startswith("DIFAL")][0]
ws_d = wb_c[difal_sheet]
ws_b = wb_b[wb_b.sheetnames[0]]

bi_hdr = list(ws_b.iter_rows(min_row=2, max_row=2, values_only=True))[0]
idx = {str(h): i for i, h in enumerate(bi_hdr) if h}

# Build BI lookup: (forn, nota, prod) -> grupo
bi_grupo = {}
for row in ws_b.iter_rows(min_row=3, values_only=True):
    if not row or not row[1]:
        continue
    key = (
        str(row[idx["Cod Fornecedor"]]).strip(),
        str(row[idx["Nota Fiscal"]]).strip().zfill(9),
        str(row[idx["Cód Produto"]]).strip(),
    )
    grupo = str(row[idx.get("Desc. Grupo Produtos", "")] or "").strip()
    bi_grupo[key] = grupo

by_grupo: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
for row in ws_d.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    key = (str(row[0]), str(row[3]), str(row[5]))
    conta = str(row[4]).strip()
    grupo = bi_grupo.get(key, "")
    if conta and conta not in ("#N/A", "#REF!"):
        by_grupo[grupo][conta] += 1

print("Grupo -> conta (top):")
for grupo in sorted(by_grupo, key=lambda g: -sum(by_grupo[g].values())):
    contas = sorted(by_grupo[grupo].items(), key=lambda x: -x[1])
    if grupo:
        print(f"  [{grupo}]: {contas[:3]}")

wb_c.close()
wb_b.close()
