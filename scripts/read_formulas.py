"""Read Excel formulas from reference DIFAL sheet."""
import openpyxl
from pathlib import Path

base = Path(__file__).resolve().parents[1]
calc = list(base.glob("*28*.xlsx"))[0]
wb = openpyxl.load_workbook(calc, read_only=False, data_only=False)
sheet = [s for s in wb.sheetnames if s.strip().startswith("DIFAL")][0]
ws = wb[sheet]

# Header row 1, find NOVO DIFAL col (X = 24)
hdr = [c.value for c in ws[1]]
print("Headers:", hdr[20:25])

# Sample rows with formulas
for row_idx in [2, 3, 50, 100]:
    row = ws[row_idx]
    print(f"\nRow {row_idx}:")
    print(f"  CONTA formula: {row[4].value}")
    print(f"  NOVO formula: {row[23].value}")
    print(f"  AJUSTE formula: {row[24].value}")

wb.close()
