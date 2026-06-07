import openpyxl
from pathlib import Path

calc = list(Path(r"c:\ProjetosIA\DIFAL_Madero").glob("*28*.xlsx"))[0]
wb = openpyxl.load_workbook(calc, read_only=True, data_only=True)

# BI QLVIEW headers
if "BI QLVIEW" in wb.sheetnames:
    ws = wb["BI QLVIEW"]
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    print("BI QLVIEW hdr count", len([h for h in hdr if h]))

# Compare BI extract vs DIFAL for multiple rows
bi_file = list(Path(r"c:\ProjetosIA\DIFAL_Madero").glob("*BI*.xlsx"))[0]
wb_bi = openpyxl.load_workbook(bi_file, read_only=True, data_only=True)
ws_bi = wb_bi[wb_bi.sheetnames[0]]
bi_hdr = list(ws_bi.iter_rows(min_row=2, max_row=2, values_only=True))[0]
idx = {h: i for i, h in enumerate(bi_hdr) if h}

difal_name = [s for s in wb.sheetnames if s.strip().startswith("DIFAL")][0]
ws_d = wb[difal_name]

matches = 0
mismatch_novo = 0
for drow in ws_d.iter_rows(min_row=2, max_row=50, values_only=True):
    if not drow[0]:
        continue
    nota_d = str(drow[3]).lstrip("0") or str(drow[3])
    forn = str(drow[0])
    prod = str(drow[5])
    novo_d = float(drow[23] or 0)
    ajuste_d = float(drow[24] or 0)
    vic = float(drow[22] or 0)

    for brow in ws_bi.iter_rows(min_row=3, values_only=True):
        if not brow[1]:
            continue
        nota_b = str(brow[idx["Nota Fiscal"]]).lstrip("0")
        if str(brow[idx["Cod Fornecedor"]]) == forn and str(brow[idx["Cód Produto"]]) == prod and nota_b == nota_d:
            v_icms_comp = float(brow[idx["Valor ICMS Complementar"]] or 0)
            d1 = float(brow[idx["D1_ICMSCOM"]] or 0)
            matches += 1
            novo_from_bi = v_icms_comp
            ajuste_from_bi = novo_from_bi - d1
            if abs(novo_d - novo_from_bi) > 0.02:
                mismatch_novo += 1
                print(f"MISMATCH novo nota {nota_d} forn {forn}: difal={novo_d} bi_vicms={novo_from_bi}")
            if abs(ajuste_d - ajuste_from_bi) > 0.02:
                print(f"MISMATCH ajuste nota {nota_d}: difal={ajuste_d} calc={ajuste_from_bi}")
            if abs(vic - d1) > 0.02:
                print(f"MISMATCH vic nota {nota_d}: difal_vic={vic} bi_d1={d1}")
            break

print(f"matched {matches}, novo mismatches {mismatch_novo}")

wb.close()
wb_bi.close()
