"""Diagnóstico novo_difal: base dupla, NCM, BI carga."""
import openpyxl
from pathlib import Path

out = Path("output/apuracao-test.xlsx")
ref = list(Path(".").glob("*28*.xlsx"))[0]
bi = list(Path(".").glob("*BI*.xlsx"))[0]


def base_dupla(vc, vi, aliq_dest_pct):
    aliq = aliq_dest_pct / 100
    base = (vc - vi) / (1 - aliq)
    return base * aliq - vi


def formula_padrao(vc, aliq_compl, aliq_icms_compl):
    return vc / ((100 - aliq_compl) / 100) * (aliq_icms_compl / 100)


def carga_simples(vc, carga_pct):
    return vc * carga_pct / 100


wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
ws = wb[[s for s in wb.sheetnames if "DIFAL" in s][0]]

print("=== Maiores |AJUSTE| ===")
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    aj = float(row[24] or 0)
    rows.append((abs(aj), row))
rows.sort(reverse=True)

for abs_aj, row in rows[:15]:
    vc, ai, vi = float(row[17]), float(row[18]), float(row[19])
    ac, aic = float(row[20]), float(row[21])
    vic, novo, aj = float(row[22]), float(row[23]), float(row[24])
    ncm = str(row[7])
    ai_pct = ai * 100 if ai <= 1 else ai
    std = formula_padrao(vc, ac, aic)
    bd = base_dupla(vc, vi, ac)
    metodo = row[25] if len(row) > 25 else ""
    ncm_r = row[26] if len(row) > 26 else ""
    carga_n = row[27] if len(row) > 27 else ""
    carga_bi = row[28] if len(row) > 28 else ""
    print(f"\n{row[0]} NF {row[3]} NCM {ncm} UF {row[1]}")
    print(f"  VC={vc:.2f} VI={vi:.2f} AI={ai_pct}% AIC={aic}%")
    print(f"  D1={vic:.2f} NOVO={novo:.2f} AJUSTE={aj:.2f}")
    print(f"  formula_padrao={std:.2f} base_dupla={bd:.2f} diff_novo_std={novo-std:.2f}")
    print(f"  audit: {metodo} ncm={ncm_r} carga_norm={carga_n} carga_bi={carga_bi}")
    if carga_n:
        cn = float(carga_n)
        print(f"  VC*carga_norm={carga_simples(vc,cn):.2f} padrao_carga_norm={formula_padrao(vc,ac,cn):.2f}")

wb.close()

# Row 2 style 85437099
print("\n=== NCM 85437099 ===")
wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
ws = wb[[s for s in wb.sheetnames if "DIFAL" in s][0]]
for row in ws.iter_rows(min_row=2, values_only=True):
    if str(row[7]) == "85437099":
        vc, vi, ac, aic = float(row[17]), float(row[19]), float(row[20]), float(row[21])
        novo, vic = float(row[23]), float(row[22])
        print(f"VC={vc} novo={novo} d1={vic} padrao={formula_padrao(vc,ac,aic):.4f} bd={base_dupla(vc,vi,ac):.4f}")
        break
wb.close()
