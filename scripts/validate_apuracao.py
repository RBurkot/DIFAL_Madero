"""Compare generated DIFAL sheet with reference."""
import openpyxl
from pathlib import Path

base = Path(__file__).resolve().parents[1]
calc = list(base.glob("*28*.xlsx"))[0]
out = base / "output" / "apuracao.xlsx"

wb_ref = openpyxl.load_workbook(calc, read_only=True, data_only=True)
wb_out = openpyxl.load_workbook(out, read_only=True, data_only=True)
ref_sheet = [s for s in wb_ref.sheetnames if s.strip().startswith("DIFAL")][0]
out_sheet = [s for s in wb_out.sheetnames if s.strip().startswith("DIFAL")][0]

ws_r = wb_ref[ref_sheet]
ws_o = wb_out[out_sheet]


def load_rows(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = (str(row[0]), str(row[3]), str(row[5]))  # forn, nota, prod
        rows.append((key, row))
    return rows


ref_rows = {k: r for k, r in load_rows(ws_r)}
out_rows = {k: r for k, r in load_rows(ws_o)}

print(f"Reference rows: {len(ref_rows)}")
print(f"Output rows: {len(out_rows)}")
print(f"Keys in both: {len(set(ref_rows) & set(out_rows))}")
print(f"Only in output: {len(set(out_rows) - set(ref_rows))}")
print(f"Only in ref: {len(set(ref_rows) - set(out_rows))}")

# Compare numeric fields for matching keys
mismatches = []
for key in set(ref_rows) & set(out_rows):
    r, o = ref_rows[key], out_rows[key]
    for idx, name in [(17, "valor_contabil"), (22, "valor_icms_compl"), (23, "novo_difal"), (24, "ajuste")]:
        rv, ov = float(r[idx] or 0), float(o[idx] or 0)
        if abs(rv - ov) > 0.02:
            mismatches.append((key, name, rv, ov, abs(rv - ov)))

print(f"Numeric mismatches (>0.02): {len(mismatches)}")
for m in mismatches[:10]:
    print(f"  {m}")

wb_ref.close()
wb_out.close()
