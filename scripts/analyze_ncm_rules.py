"""Analisa Planilha1 e impacto no cálculo novo_difal."""
import re
from pathlib import Path

import openpyxl

base = Path(__file__).resolve().parents[1]
ref = list(base.glob("*28*.xlsx"))[0]
gerado = base / "dist" / "output" / "resultado-052026.xlsx"
bi = list(base.glob("*BI*.xlsx"))[0]


def parse_planilha1(path: Path) -> dict[str, dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Planilha1"]
    rules: dict[str, dict] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        ncm = str(row[0]).strip()
        texto = str(row[1] or "").strip()
        carga = None
        if texto:
            m = re.search(r"(\d+[,.]?\d*)\s*%", texto)
            if m:
                carga = float(m.group(1).replace(",", "."))
        rules[ncm] = {
            "tem_regra": bool(texto),
            "texto": texto,
            "carga_pct": carga,
            "ufs": ["RS", "SC", "PR", "SP", "RJ", "MG"] if "Sul e Sudeste" in texto else None,
        }
    wb.close()
    return rules


def difal_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = [s for s in wb.sheetnames if s.strip().startswith("DIFAL")][0]
    rows = []
    for row in wb[sheet].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        vc = float(row[17] or 0)
        ac = float(row[20] or 0)
        aic = float(row[21] or 0)
        novo = float(row[23] or 0)
        std = vc / ((100 - ac) / 100) * (aic / 100)
        rows.append({
            "key": (str(row[0]), str(row[3]), str(row[5])),
            "ncm": str(row[7]).strip(),
            "uf": str(row[1]).strip(),
            "vc": vc,
            "ac": ac,
            "aic": aic,
            "novo": novo,
            "std": std,
            "usa_std": abs(novo - std) < 0.02,
        })
    wb.close()
    return rows


def bi_carga_by_key(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    idx = {str(h): i for i, h in enumerate(hdr) if h}
    out = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[idx["Cod Fornecedor"]]:
            continue
        k = (
            str(row[idx["Cod Fornecedor"]]).strip(),
            str(row[idx["Nota Fiscal"]]).strip().zfill(9),
            str(row[idx["Cód Produto"]]).strip(),
        )
        carga = row[idx.get("Carga Efetiva DIFAL", -1)] if "Carga Efetiva DIFAL" in idx else None
        out[k] = {
            "ncm": str(row[idx["NCM"]]).strip(),
            "uf": str(row[idx["Estado"]]).strip(),
            "carga_bi": float(carga) if carga not in (None, "", 0) else None,
        }
    wb.close()
    return out


rules = parse_planilha1(ref)
special = {n: r for n, r in rules.items() if r["tem_regra"]}
print("=== NCMs com regra dissertativa (Planilha1) ===")
for ncm, r in special.items():
    print(f"  {ncm}: carga={r['carga_pct']}% ufs={r['ufs']}")

ref_rows = difal_rows(ref)
out_rows = difal_rows(gerado) if gerado.exists() else []
bi_map = bi_carga_by_key(bi)

print("\n=== Uso na DIFAL referência (NCMs especiais) ===")
for ncm in special:
    subset = [r for r in ref_rows if r["ncm"] == ncm]
    if not subset:
        print(f"  {ncm}: 0 linhas")
        continue
    std_ok = sum(1 for r in subset if r["usa_std"])
    carga = special[ncm]["carga_pct"]
    carga_ok = 0
    if carga:
        for r in subset:
            if abs(r["novo"] - r["vc"] * carga / 100) < 0.02:
                carga_ok += 1
    print(f"  {ncm}: {len(subset)} linhas | formula_std={std_ok} | formula_carga_txt={carga_ok}")

print("\n=== BI: Carga Efetiva DIFAL preenchida (NCMs especiais) ===")
for ncm in special:
    cargas = [
        bi_map[k]["carga_bi"]
        for k, v in bi_map.items()
        if v["ncm"] == ncm and v["carga_bi"]
    ]
    print(f"  {ncm}: {len(cargas)} linhas BI com carga>0 (amostra: {cargas[:3]})")

# Mismatches gerado vs ref for special NCMs
if out_rows:
    ref_by_key = {r["key"]: r for r in ref_rows}
    print("\n=== Divergencias gerado vs ref (NCMs especiais, chaves em comum) ===")
    for r in out_rows:
        if r["ncm"] not in special:
            continue
        ref_r = ref_by_key.get(r["key"])
        if not ref_r:
            continue
        if abs(r["novo"] - ref_r["novo"]) > 0.02:
            bi = bi_map.get(r["key"], {})
            print(f"  {r['key']} ncm={r['ncm']} uf={r['uf']}")
            print(f"    ref_novo={ref_r['novo']:.2f} out_novo={r['novo']:.2f} bi_carga={bi.get('carga_bi')}")
