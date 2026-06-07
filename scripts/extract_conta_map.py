"""Extract conta contabil mapping from reference DIFAL sheet."""
import openpyxl
from pathlib import Path
from collections import defaultdict

base = Path(__file__).resolve().parents[1]
calc = list(base.glob("*28*.xlsx"))[0]
wb = openpyxl.load_workbook(calc, read_only=True, data_only=True)
difal_sheet = [s for s in wb.sheetnames if s.strip().startswith("DIFAL")][0]
ws = wb[difal_sheet]

# CFOP -> conta counts, also grupo if available
by_cfop: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
by_grupo: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    cfop = str(row[8]).strip()
    conta = str(row[4]).strip()
    if conta and conta not in ("#N/A", "#REF!", "*"):
        by_cfop[cfop][conta] += 1

print(f"Sheet: {difal_sheet}")
for cfop in sorted(by_cfop):
    contas = sorted(by_cfop[cfop].items(), key=lambda x: -x[1])
    print(f"CFOP {cfop}: {contas[:5]}")

# Count #N/A
na = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row and row[4] == "#N/A")
total = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0])
print(f"\n#N/A contas: {na}/{total}")

wb.close()
