"""Temporary analysis of reference spreadsheets."""
import openpyxl
from pathlib import Path

base = Path(r"c:\ProjetosIA\DIFAL_Madero")
calc = list(base.glob("*28*.xlsx"))[0]
bi = list(base.glob("*BI*.xlsx"))[0]

wb_c = openpyxl.load_workbook(calc, read_only=True, data_only=True)
wb_b = openpyxl.load_workbook(bi, read_only=True, data_only=True)

# BI full header row 1
ws_b = wb_b[wb_b.sheetnames[0]]
bi_hdr = list(ws_b.iter_rows(min_row=2, max_row=2, values_only=True))[0]
print("BI columns:", len(bi_hdr))
for i, h in enumerate(bi_hdr):
    if h:
        print(f"  {i}: {h}")

# Map key BI indices
idx = {str(h): i for i, h in enumerate(bi_hdr) if h}

difal_sheet = [s for s in wb_c.sheetnames if s.strip().startswith("DIFAL")][0]
ws_d = wb_c[difal_sheet]

# Compare first matching line SP nota 13
target_nota = "000000013"
target_forn = "221235"
target_prod = "10724305001300"

bi_row = None
for row in ws_b.iter_rows(min_row=3, values_only=True):
    if not row or not row[1]:
        continue
    nota = str(row[idx.get("Nota Fiscal", 9)]).zfill(9) if row[idx.get("Nota Fiscal", 9)] else ""
    forn = str(row[idx.get("Cod Fornecedor", 1)])
    prod = str(row[idx.get("Cód Produto", 10)])
    if "13" in nota and forn == target_forn and prod == target_prod:
        bi_row = row
        break

difal_row = None
for row in ws_d.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    if str(row[0]) == target_forn and str(row[3]) == target_nota and str(row[5]) == target_prod:
        difal_row = row
        break

print("\n--- BI row sample (nota 13) ---")
if bi_row:
    print("Valor Contábil:", bi_row[idx["Valor Contábil"]])
    print("Alíquota ICMS:", bi_row[idx["Alíquota ICMS"]])
    print("Alíquota Compl.:", bi_row[idx["Alíquota Compl."]])
    print("Aliquota ICMS Complementar:", bi_row[idx["Aliquota ICMS Complementar"]])
    print("Valor ICMS Complementar:", bi_row[idx.get("Valor ICMS Complementar", idx.get("D1_ICMSCOM", 31))])
    # find D1_ICMSCOM
    for k in idx:
        if "ICMS" in k or "Compl" in k or "DIFAL" in k or "Carga" in k:
            print(f"  {k}:", bi_row[idx[k]])

print("\n--- DIFAL row sample ---")
if difal_row:
    print("VALOR CONTÁBIL:", difal_row[17])
    print("ALÍQUOTA ICMS:", difal_row[18])
    print("ALÍQUOTA COMPLEMENTAR:", difal_row[20])
    print("ALIQUOTA ICMS COMPLEMENTAR:", difal_row[21])
    print("VALOR ICMS COMPLEMENTAR:", difal_row[22])
    print("NOVO DIFAL:", difal_row[23])
    print("AJUSTE:", difal_row[24])
    vc = float(difal_row[17])
    aliq_comp = float(difal_row[21]) if difal_row[21] else 0
    novo = vc * aliq_comp / 100
    print("calc novo (vc*aliq/100):", novo)
    print("ajuste calc:", novo - float(difal_row[22]))

# NCM rules sheet
if "Planilha1" in wb_c.sheetnames:
    ws_p = wb_c["Planilha1"]
    print("\n--- NCM rules ---")
    for row in ws_p.iter_rows(max_row=10, values_only=True):
        if row[0]:
            print(row)

wb_c.close()
wb_b.close()
