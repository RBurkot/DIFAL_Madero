"""Diagnóstico: diferença de linhas DIFAL gerada vs referência."""
import openpyxl
from collections import Counter
from datetime import datetime
from pathlib import Path

base = Path(__file__).resolve().parents[1]
gerado = base / "dist" / "output" / "resultado-052026.xlsx"
ref_wb = list(base.glob("*28*.xlsx"))[0]
bi_wb = list(base.glob("*BI*.xlsx"))[0]


def load_difal(path: Path) -> tuple[str, list[tuple]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = [s for s in wb.sheetnames if s.strip().upper().startswith("DIFAL")][0]
    ws = wb[sheet]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = (str(row[0]).strip(), str(row[3]).strip(), str(row[5]).strip())
        rows.append((key, row))
    wb.close()
    return sheet, rows


def key_set(rows):
    return {k for k, _ in rows}


print("=== CONTAGENS ===")
ref_sheet, ref_rows = load_difal(ref_wb)
out_sheet, out_rows = load_difal(gerado)
print(f"Referência [{ref_sheet}]: {len(ref_rows)} linhas")
print(f"Gerado     [{out_sheet}]: {len(out_rows)} linhas")
print(f"Diferença: {len(out_rows) - len(ref_rows):+d}")

ref_keys = key_set(ref_rows)
out_keys = key_set(out_rows)
ref_by_key = {}
for k, r in ref_rows:
    ref_by_key.setdefault(k, []).append(r)
out_by_key = {}
for k, r in out_rows:
    out_by_key.setdefault(k, []).append(r)

only_out = out_keys - ref_keys
only_ref = ref_keys - out_keys
both = ref_keys & out_keys
print(f"\nChaves únicas ref: {len(ref_keys)} | gerado: {len(out_keys)}")
print(f"Só no gerado: {len(only_out)} | Só na ref: {len(only_ref)} | Em ambos: {len(both)}")

# CFOP das extras
cfop_extra = Counter()
for k in only_out:
    for r in out_by_key[k]:
        cfop_extra[str(r[8])] += 1
print(f"\nCFOPs das {len(only_out)} chaves só no gerado: {dict(cfop_extra)}")

# BI analysis for extra keys
wb_b = openpyxl.load_workbook(bi_wb, read_only=True, data_only=True)
ws_b = wb_b[wb_b.sheetnames[0]]
hdr = list(ws_b.iter_rows(min_row=2, max_row=2, values_only=True))[0]
idx = {str(h): i for i, h in enumerate(hdr) if h}

extra_bi = []
for row in ws_b.iter_rows(min_row=3, values_only=True):
    if not row or not row[idx["Cod Fornecedor"]]:
        continue
    k = (
        str(row[idx["Cod Fornecedor"]]).strip(),
        str(row[idx["Nota Fiscal"]]).strip().zfill(9),
        str(row[idx["Cód Produto"]]).strip(),
    )
    if k in only_out:
        mes = row[idx["MES_ANO_ENTRADA"]]
        extra_bi.append({
            "key": k,
            "cfop": str(row[idx["Cód Fiscal"]]).strip(),
            "tes": str(row[idx.get("Descr. TES", "")] or ""),
            "grupo": str(row[idx.get("Desc. Grupo Produtos", "")] or ""),
            "mes": mes,
            "dia": mes.day if isinstance(mes, datetime) else None,
            "uf": str(row[idx["Estado"]]).strip(),
        })
wb_b.close()

print(f"\nLinhas BI correspondentes às chaves extras: {len(extra_bi)}")
if extra_bi:
    tes_c = Counter(x["tes"] for x in extra_bi)
    grupo_c = Counter(x["grupo"] for x in extra_bi)
    dia_c = Counter(x["dia"] for x in extra_bi if x["dia"])
    print("TES mais frequentes nas extras:")
    for t, n in tes_c.most_common(8):
        print(f"  [{n}] {t[:60]}")
    print("Grupos nas extras:")
    for g, n in grupo_c.most_common(8):
        print(f"  [{n}] {g}")
    print("Dias (MES_ANO_ENTRADA) nas extras:", dict(sorted(dia_c.items())))

# BI QLVIEW in reference - keys present?
wb_r = openpyxl.load_workbook(ref_wb, read_only=True, data_only=True)
if "BI QLVIEW" in wb_r.sheetnames:
    ws_q = wb_r["BI QLVIEW"]
    qlview_keys = set()
    for row in ws_q.iter_rows(min_row=3, values_only=True):
        if not row or not row[1]:
            continue
        # QLVIEW layout: col1=forn, col9=nota?, col10=prod - from prior analysis
        qlview_keys.add((str(row[1]).strip(), str(row[9]).strip().zfill(9), str(row[10]).strip()))
    in_qlview = sum(1 for k in only_out if k in qlview_keys)
    not_in_qlview = len(only_out) - in_qlview
    print(f"\nChaves só no gerado presentes em BI QLVIEW (ref): {in_qlview}/{len(only_out)}")
    print(f"Chaves só no gerado AUSENTES em BI QLVIEW (ref): {not_in_qlview}/{len(only_out)}")
    # ref keys all in qlview?
    ref_in_q = sum(1 for k in ref_keys if k in qlview_keys)
    print(f"Chaves da ref DIFAL presentes em BI QLVIEW: {ref_in_q}/{len(ref_keys)}")
wb_r.close()

# Duplicate keys
dup_ref = sum(1 for k, v in ref_by_key.items() if len(v) > 1)
dup_out = sum(1 for k, v in out_by_key.items() if len(v) > 1)
print(f"\nChaves duplicadas: ref={dup_ref} grupos | gerado={dup_out} grupos")

print("\n=== AMOSTRA chaves só no gerado (5) ===")
for k in list(only_out)[:5]:
    bi = next((x for x in extra_bi if x["key"] == k), None)
    print(k, bi)
