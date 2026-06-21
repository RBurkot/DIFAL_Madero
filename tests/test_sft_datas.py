from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

from difal_importacao.reader import build_chave_rastreio, read_auxiliar_sft
from difal_importacao.config import ImportacaoConfig
from difal_importacao.models import LinhaApuracaoDifal, PeriodoApuracao
from difal_importacao.transformer import transform_linha

# SFT: Chave col A (=K&G&Q); INDUSTRIA-IMPORTAÇÃO: Chave col K
SFT_CHAVE_COL = column_index_from_string("A") - 1
SFT_PART_K = column_index_from_string("K") - 1
SFT_PART_G = column_index_from_string("G") - 1
SFT_PART_Q = column_index_from_string("Q") - 1
NOME_COL = column_index_from_string("I") - 1
ROW_WIDTH = column_index_from_string("Q")


def _write_sft_sheet(path: Path, chave: str, emissao: datetime, entrada: datetime) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SFT "
    headers = [None] * ROW_WIDTH
    headers[SFT_CHAVE_COL] = "Chave"
    headers[1] = "Nota Fiscal"
    headers[2] = "Cd Produto"
    headers[3] = "Cd. Forn"
    headers[4] = "Lj. Forn"
    headers[5] = "Dt Emissao"
    headers[6] = "Dt Entrada"
    headers[NOME_COL] = "Nm Forn"
    ws.append(headers)

    row = [None] * ROW_WIDTH
    row[SFT_CHAVE_COL] = chave
    row[1] = "000000100"
    row[2] = "P001"
    row[3] = "123"
    row[SFT_PART_K] = 100
    row[SFT_PART_G] = "123"
    row[SFT_PART_Q] = "P001"
    row[NOME_COL] = "Fornecedor SFT Coluna I"
    row[4] = "0001"
    row[5] = emissao
    row[6] = entrada
    ws.append(row)
    wb.save(path)
    wb.close()


def test_sft_indexa_por_coluna_a_e_retorna_datas(tmp_path):
    emissao = datetime(2026, 5, 10)
    entrada = datetime(2026, 5, 15)
    chave = "100123P001"
    path = tmp_path / "ref.xlsx"
    _write_sft_sheet(path, chave, emissao, entrada)

    lookups = read_auxiliar_sft(path)

    assert chave in lookups
    assert lookups[chave].data_emissao == emissao
    assert lookups[chave].data_entrada == entrada
    assert lookups[chave].nome_fornecedor == "Fornecedor SFT Coluna I"


def test_transform_preenche_colunas_l_m_via_chave_sft(tmp_path):
    emissao = datetime(2026, 5, 10)
    entrada = datetime(2026, 5, 15)
    chave = build_chave_rastreio("000000100", "123", "P001")
    path = tmp_path / "ref.xlsx"
    _write_sft_sheet(path, chave, emissao, entrada)
    lookups = read_auxiliar_sft(path)

    linha = LinhaApuracaoDifal(
        fornecedor_cod="123",
        uf_origem="SP",
        filial_cod="01GDIN0004",
        nota_fiscal="000000100",
        conta_contabil="11010020003",
        cod_produto="P001",
        produto_desc="Produto",
        ncm="",
        cfop="2556",
        valor_contabil=100.0,
        valor_icms_complementar=10.0,
        novo_difal=12.0,
        ajuste=2.0,
    )
    periodo = PeriodoApuracao(5, 2026, "05.2026", "DIFAL 05.2026 ")

    lanc = transform_linha(
        linha,
        periodo,
        ImportacaoConfig(),
        lookups=lookups,
        sb1_map={"P001": "11010020003"},
    )

    assert lanc is not None
    assert lanc.chave == chave
    assert lanc.data_emissao == emissao
    assert lanc.data_entrada == entrada
    assert lanc.nome_fornecedor == "Fornecedor SFT Coluna I"


def test_sft_pesquisa_por_chave_a_sem_cdforn(tmp_path):
    emissao = datetime(2026, 5, 20)
    entrada = datetime(2026, 5, 21)
    chave = "100123P001"
    path = tmp_path / "ref_sem_forn.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SFT "
    headers = [None] * (NOME_COL + 1)
    headers[SFT_CHAVE_COL] = "Chave"
    headers[NOME_COL] = "Nm Forn"
    headers[5] = "Dt Emissao"
    headers[6] = "Dt Entrada"
    ws.append(headers)
    row = [None] * (NOME_COL + 1)
    row[SFT_CHAVE_COL] = chave
    row[NOME_COL] = "Fornecedor via chave"
    row[5] = emissao
    row[6] = entrada
    ws.append(row)
    wb.save(path)
    wb.close()

    lookups = read_auxiliar_sft(path)

    assert chave in lookups
    assert lookups[chave].data_emissao == emissao
    assert lookups[chave].data_entrada == entrada
    assert lookups[chave].nome_fornecedor == "Fornecedor via chave"


def test_sft_nome_fornecedor_coluna_i_mesmo_com_fornecedor_generico(tmp_path):
    chave = "100123P001"
    path = tmp_path / "ref_nome_i.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SFT "
    headers = [None] * (NOME_COL + 1)
    headers[SFT_CHAVE_COL] = "Chave"
    headers[3] = "Fornecedor"  # coluna D — código, não nome
    headers[NOME_COL] = "Nm Forn"
    headers[5] = "Dt Emissao"
    headers[6] = "Dt Entrada"
    ws.append(headers)

    row = [None] * (NOME_COL + 1)
    row[SFT_CHAVE_COL] = chave
    row[3] = "123"
    row[NOME_COL] = "RAZAO SOCIAL CORRETA"
    row[5] = datetime(2026, 5, 10)
    row[6] = datetime(2026, 5, 15)
    ws.append(row)
    wb.save(path)
    wb.close()

    lookups = read_auxiliar_sft(path)

    assert chave in lookups
    assert lookups[chave].nome_fornecedor == "RAZAO SOCIAL CORRETA"


def test_chave_ignora_sufixo_excel_em_fornecedor_e_produto():
    base = build_chave_rastreio("000000100", "123", "P001")
    assert build_chave_rastreio("000000100", "123.0", "P001") == base
    assert build_chave_rastreio("000000100", "123", "P001.0") == base


def test_sft_col_a_formula_sem_cache_usa_k_g_q(tmp_path):
    """Col A =K&G&Q sem cache Excel → indexa pelas colunas K, G e Q."""
    emissao = datetime(2026, 5, 10)
    entrada = datetime(2026, 5, 15)
    chave = build_chave_rastreio("000000100", "123", "P001")
    path = tmp_path / "ref_formula.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SFT "
    headers = [None] * ROW_WIDTH
    headers[SFT_CHAVE_COL] = "Chave"
    headers[11] = "Dt Emissao"
    headers[12] = "Dt Entrada"
    headers[NOME_COL] = "Nm Forn"
    ws.append(headers)

    ws.append([None] * ROW_WIDTH)
    ws["A2"] = "=K2&G2&Q2"
    ws["K2"] = 100
    ws["G2"] = "123"
    ws["Q2"] = "P001"
    ws["I2"] = "Fornecedor Formula"
    ws["L2"] = emissao
    ws["M2"] = entrada
    wb.save(path)
    wb.close()

    lookups = read_auxiliar_sft(path)

    assert chave in lookups
    assert lookups[chave].nome_fornecedor == "Fornecedor Formula"
    assert lookups[chave].data_emissao == emissao
    assert lookups[chave].data_entrada == entrada


def test_sft_chave_real_k_g_q_usuario(tmp_path):
    """Chave informada pelo usuário: 60690520106420904435025500 (= K6069052 & G010642090 & Q4435025500)."""
    chave = "60690520106420904435025500"
    path = tmp_path / "ref_chave_usuario.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SFT "
    headers = [None] * ROW_WIDTH
    headers[SFT_CHAVE_COL] = "Chave"
    headers[NOME_COL] = "Nm Forn"
    ws.append(headers)

    ws.append([None] * ROW_WIDTH)
    ws["A2"] = "=K2&G2&Q2"
    ws["K2"] = 6069052
    ws["G2"] = "010642090"
    ws["Q2"] = 4435025500
    ws["I2"] = "Fornecedor Teste Real"
    wb.save(path)
    wb.close()

    assert build_chave_rastreio("6069052", "010642090", "4435025500") == chave

    lookups = read_auxiliar_sft(path)

    assert chave in lookups
    assert lookups[chave].nome_fornecedor == "Fornecedor Teste Real"


def test_chave_from_sft_cell_ignora_texto_formula():
    from difal_importacao.reader import chave_from_sft_cell

    assert chave_from_sft_cell("=TEXT(B2,\"0\")&D2&C2") == ""
    assert chave_from_sft_cell(None) == ""
    assert chave_from_sft_cell("100123P001") == "100123P001"
    assert chave_from_sft_cell(100123001) == "100123001"


def test_sft_cabecalho_linha2_e_data_com_acento(tmp_path):
    chave = "100123P001"
    path = tmp_path / "ref_header2.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SFT "
    ws.append(["metadado"])
    headers = [None] * 13
    headers[SFT_CHAVE_COL] = "Chave"
    headers[NOME_COL] = "Nm Forn"
    headers[11] = "Dt Emissão"  # L
    headers[12] = "Dt Entrada"  # M
    ws.append(headers)

    row = [None] * 13
    row[SFT_CHAVE_COL] = chave
    row[NOME_COL] = "Fornecedor Header2"
    row[11] = "21/06/2026"
    row[12] = 46203  # serial Excel
    ws.append(row)
    wb.save(path)
    wb.close()

    lookups = read_auxiliar_sft(path)

    assert chave in lookups
    assert lookups[chave].nome_fornecedor == "Fornecedor Header2"
    assert lookups[chave].data_emissao is not None
    assert lookups[chave].data_entrada is not None
