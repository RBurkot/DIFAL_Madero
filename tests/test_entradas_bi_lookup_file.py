import openpyxl

from difal_importacao.entradas_bi_lookup import (
    COD_ITEM_COL,
    load_entradas_item_contabil,
)


def test_load_entradas_item_contabil_coluna_ak(tmp_path):
    path = tmp_path / "ref.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ENTRADAS-BI"
    ws.append(["x"] * (COD_ITEM_COL + 1))
    ws.cell(row=1, column=COD_ITEM_COL + 1, value="COD ITEM")
    ws.cell(row=1, column=11, value="Cód Produto")
    ws.cell(row=2, column=11, value="P001")
    ws.cell(row=2, column=COD_ITEM_COL + 1, value="AK123")
    ws.cell(row=3, column=11, value="P002")
    ws.cell(row=3, column=COD_ITEM_COL + 1, value=456)
    wb.save(path)
    wb.close()

    mapping = load_entradas_item_contabil(str(path))

    assert mapping["P001"] == "AK123"
    assert mapping["P002"] == "456"
    load_entradas_item_contabil.cache_clear()
