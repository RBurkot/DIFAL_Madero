import json
from pathlib import Path

import openpyxl

from difal_importacao.config import ImportacaoConfig
from difal_importacao.models import RelatorioReconciliacao
from difal_importacao.reader import build_chave_rastreio


def _read_ref_lancamentos(path: Path) -> dict[str, dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["INDUSTRIA-IMPORTAÇÃO"]
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and row[0] == "LOJA":
            header_row = i
            break
    if header_row is None:
        wb.close()
        return {}

    ref: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
        if not row or not row[0] or str(row[0]) not in ("0001", "0002"):
            continue
        chave = str(row[10])
        ref[chave] = {
            "debito": str(row[14]),
            "credito": str(row[18]),
            "valor": float(row[22] or 0),
        }
    wb.close()
    return ref


def reconciliar(
    gerado_path: Path,
    referencia_path: Path,
    config: ImportacaoConfig,
    periodo: str,
) -> RelatorioReconciliacao:
    ref = _read_ref_lancamentos(referencia_path)
    wb = openpyxl.load_workbook(gerado_path, read_only=True, data_only=True)
    ws = wb["INDUSTRIA-IMPORTAÇÃO"]
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and row[0] == "LOJA":
            header_row = i
            break

    divergencias = []
    gerados = 0
    for row in ws.iter_rows(min_row=(header_row or 0) + 2, values_only=True):
        if not row or not row[0] or str(row[0]) not in ("0001", "0002"):
            continue
        gerados += 1
        chave = str(row[10])
        if chave not in ref:
            divergencias.append({"chave": chave, "tipo": "extra_gerada"})
            continue
        exp = ref[chave]
        val = float(row[22] or 0)
        if abs(val - exp["valor"]) > config.tolerancia_linha:
            divergencias.append({
                "chave": chave, "campo": "valor",
                "esperado": exp["valor"], "obtido": val,
            })

    wb.close()
    ausentes = [k for k in ref if k not in {d.get("chave") for d in divergencias if "extra" not in d.get("tipo", "")}]
    resultado = "APROVADO" if not divergencias else "REPROVADO"
    return RelatorioReconciliacao(
        periodo=periodo,
        lancamentos_gerados=gerados,
        divergencias=divergencias,
        resultado=resultado,
    )


def save_relatorio(rel: RelatorioReconciliacao, path: Path) -> None:
    path.write_text(
        json.dumps(rel.__dict__, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )
