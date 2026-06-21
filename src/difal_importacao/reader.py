import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.utils.datetime import from_excel

from difal_apuracao.sb1_lookup import normalize_cfop, normalize_codigo
from difal_importacao.models import FornecedorLookup, LinhaApuracaoDifal, PeriodoApuracao

# INDUSTRIA-IMPORTAÇÃO col K (Chave) → lookup na SFT col A (= K & G & Q da própria SFT)
SFT_CHAVE_COL = column_index_from_string("A") - 1
SFT_CHAVE_PART_K = column_index_from_string("K") - 1
SFT_CHAVE_PART_G = column_index_from_string("G") - 1
SFT_CHAVE_PART_Q = column_index_from_string("Q") - 1
SFT_NOME_COL = column_index_from_string("I") - 1  # coluna I — Nome Fornecedor
SFT_DATA_EMISSAO_COL = column_index_from_string("L") - 1  # fallback coluna L
SFT_DATA_ENTRADA_COL = column_index_from_string("M") - 1  # fallback coluna M

DIFAL_SHEET_RE = re.compile(r"^DIFAL\s+(\d{2})\.(\d{4})\s*$", re.IGNORECASE)

DIFAL_COLUMNS = {
    "FORNECEDOR": "fornecedor_cod",
    "ESTADO": "uf_origem",
    "COD FILIAL": "filial_cod",
    "NOTA FISCAL": "nota_fiscal",
    "CONTA CONTABIL": "conta_contabil",
    "COD PRODUTO": "cod_produto",
    "PRODUTO": "produto_desc",
    "NCM": "ncm",
    "COD FISCAL": "cfop",
    "VALOR CONTÁBIL": "valor_contabil",
    "VALOR ICMS COMPLEMENTAR": "valor_icms_complementar",
    "NOVO DIFAL": "novo_difal",
    "AJUSTE": "ajuste",
}


class DIFALValidationError(Exception):
    pass


def _num(val, default=0.0) -> float:
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def find_difal_sheet(wb) -> tuple[str, PeriodoApuracao]:
    for name in wb.sheetnames:
        m = DIFAL_SHEET_RE.match(name.strip())
        if m:
            mes, ano = int(m.group(1)), int(m.group(2))
            return name, PeriodoApuracao(mes=mes, ano=ano, label=f"{mes:02d}.{ano}", aba_origem=name)
    raise DIFALValidationError("Aba DIFAL MM.AAAA não encontrada")


def read_difal(path: Path | str, aba: str | None = None) -> tuple[PeriodoApuracao, list[LinhaApuracaoDifal]]:
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if aba:
        if aba not in wb.sheetnames:
            wb.close()
            raise DIFALValidationError(f"Aba '{aba}' não encontrada")
        m = DIFAL_SHEET_RE.match(aba.strip())
        if not m:
            raise DIFALValidationError(f"Aba '{aba}' não corresponde ao padrão DIFAL MM.AAAA")
        periodo = PeriodoApuracao(int(m.group(1)), int(m.group(2)), f"{int(m.group(1)):02d}.{m.group(2)}", aba)
        sheet_name = aba
    else:
        sheet_name, periodo = find_difal_sheet(wb)

    ws = wb[sheet_name]
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_idx = {str(h).strip(): i for i, h in enumerate(hdr) if h}
    missing = [c for c in DIFAL_COLUMNS if c not in col_idx]
    if missing:
        wb.close()
        raise DIFALValidationError(f"Colunas DIFAL ausentes: {', '.join(missing)}")

    linhas: list[LinhaApuracaoDifal] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col_idx["FORNECEDOR"]]:
            continue
        linhas.append(
            LinhaApuracaoDifal(
                fornecedor_cod=normalize_codigo(row[col_idx["FORNECEDOR"]]),
                uf_origem=str(row[col_idx["ESTADO"]]).strip(),
                filial_cod=str(row[col_idx["COD FILIAL"]]).strip(),
                nota_fiscal=str(row[col_idx["NOTA FISCAL"]]).strip(),
                conta_contabil=str(row[col_idx["CONTA CONTABIL"]]).strip(),
                cod_produto=normalize_codigo(row[col_idx["COD PRODUTO"]]),
                produto_desc=str(row[col_idx.get("PRODUTO", col_idx["COD PRODUTO"])] or "").strip(),
                ncm=str(row[col_idx.get("NCM", 0)] or "").strip() if "NCM" in col_idx else "",
                cfop=normalize_cfop(row[col_idx["COD FISCAL"]]),
                valor_contabil=_num(row[col_idx["VALOR CONTÁBIL"]]),
                valor_icms_complementar=_num(row[col_idx["VALOR ICMS COMPLEMENTAR"]]),
                novo_difal=_num(row[col_idx["NOVO DIFAL"]]),
                ajuste=_num(row[col_idx["AJUSTE"]]),
            )
        )
    wb.close()
    return periodo, linhas


def build_chave_rastreio(nota: str, fornecedor: str, produto: str) -> str:
    nota_int = str(int(str(nota).lstrip("0") or "0"))
    forn = normalize_codigo(fornecedor)
    prod = normalize_codigo(produto)
    return f"{nota_int}{forn}{prod}"


def normalize_chave(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        if val == int(val):
            return str(int(val))
        return str(val).strip()
    chave = str(val).strip()
    if chave.endswith(".0") and chave[:-2].isdigit():
        return chave[:-2]
    return chave


def chave_from_sft_cell(val) -> str:
    """Valor da col A (Chave). Ignora fórmulas — usa só resultado materializado pelo Excel."""
    if val is None:
        return ""
    if isinstance(val, str) and val.lstrip().startswith("="):
        return ""
    return normalize_chave(val)


def _part_chave(val) -> str:
    """Parte da chave SFT (=K&G&Q), espelhando concatenação Excel &."""
    if val is None:
        return ""
    if isinstance(val, str) and val.lstrip().startswith("="):
        return ""
    if isinstance(val, (int, float)):
        if val == int(val):
            return str(int(val))
        return str(val).strip()
    text = str(val).strip()
    if text.endswith(".0"):
        head = text[:-2]
        if head:
            return head
    return text


def build_chave_sft_row(cells: list) -> str:
    """Replica =K2&G2&Q2 da aba SFT (colunas fixas K, G, Q)."""
    parts: list[str] = []
    for col in (SFT_CHAVE_PART_K, SFT_CHAVE_PART_G, SFT_CHAVE_PART_Q):
        if col < len(cells):
            parts.append(_part_chave(cells[col]))
        else:
            parts.append("")
    chave = "".join(parts)
    return chave if any(parts) else ""


def _sft_component_col(
    idx: dict[str, int],
    idx_norm: dict[str, int],
    candidates: tuple[str, ...],
) -> int | None:
    return _header_col(idx, idx_norm, candidates)


def _parse_date(val) -> datetime | None:
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        try:
            dt = from_excel(val)
            if isinstance(dt, datetime):
                return dt
            if isinstance(dt, date):
                return datetime(dt.year, dt.month, dt.day)
        except (TypeError, ValueError):
            return None
    if isinstance(val, str):
        text = val.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


def _normalize_header(text: str) -> str:
    t = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", t).strip().lower()


def _build_header_index(hdr: tuple) -> tuple[dict[str, int], dict[str, int]]:
    idx: dict[str, int] = {}
    idx_norm: dict[str, int] = {}
    for i, h in enumerate(hdr):
        if h is None:
            continue
        raw = str(h).strip()
        idx[raw] = i
        idx_norm[_normalize_header(raw)] = i
    return idx, idx_norm


def _header_col(idx: dict[str, int], idx_norm: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    for c in candidates:
        if c in idx:
            return idx[c]
    for c in candidates:
        c_norm = _normalize_header(c)
        if c_norm in idx_norm:
            return idx_norm[c_norm]
    return None


def _chave_col_index(idx: dict[str, int], idx_norm: dict[str, int]) -> int:
    chave_col = _header_col(idx, idx_norm, ("Chave", "CHAVE"))
    if chave_col is not None:
        return chave_col
    return SFT_CHAVE_COL


def lookup_fornecedor_by_chave(
    lookups: dict[str, FornecedorLookup] | None,
    chave: str,
) -> FornecedorLookup | None:
    if not lookups:
        return None
    for key in (chave, normalize_chave(chave)):
        if key and key in lookups:
            return lookups[key]
    return None


def _nome_forn_col_index(idx: dict[str, int], idx_norm: dict[str, int]) -> int:
    # INDUSTRIA-IMPORTAÇÃO col C ← SFT col I (Nm Forn)
    nome_col = _header_col(
        idx,
        idx_norm,
        ("Nm Forn", "Nm Fornecedor", "Nome Fornecedor", "Nome Forne", "NOME FORNE"),
    )
    if nome_col is not None:
        return nome_col
    return SFT_NOME_COL


def _date_col_index(idx: dict[str, int], idx_norm: dict[str, int], *, emissao: bool) -> int:
    if emissao:
        col = _header_col(idx, idx_norm, ("Dt Emissao", "Dt Emissão", "Data Emissao", "Data Emissão"))
        if col is not None:
            return col
        return SFT_DATA_EMISSAO_COL
    col = _header_col(idx, idx_norm, ("Dt Entrada", "Data Entrada"))
    if col is not None:
        return col
    return SFT_DATA_ENTRADA_COL


def _find_sft_header(ws) -> tuple[int, tuple]:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        if not row:
            continue
        row_list = [str(c).strip() if c is not None else "" for c in row]
        row_norm = {_normalize_header(c) for c in row_list if c}
        if "chave" in row_norm or "cd. forn" in row_norm or "cd forn" in row_norm:
            return row_idx, row
    # fallback layout antigo: cabeçalho primeira linha
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return 1, hdr


def _cell_str(cells: list, col_idx: int | None) -> str:
    if col_idx is None or col_idx >= len(cells) or cells[col_idx] is None:
        return ""
    return str(cells[col_idx]).strip()


def load_sft_lookups(
    source: Path | str | None = None,
    referencia: Path | str | None = None,
    aba: str = "SFT ",
) -> dict[str, FornecedorLookup]:
    """Carrega SFT do workbook de entrada e mescla a referência (*28*.xlsx), que prevalece."""
    lookups: dict[str, FornecedorLookup] = {}
    if source and Path(source).exists():
        lookups.update(read_auxiliar_sft(source, aba))
    if referencia and Path(referencia).exists():
        ref_path = Path(referencia)
        if not source or ref_path.resolve() != Path(source).resolve():
            lookups.update(read_auxiliar_sft(referencia, aba))
    return lookups


def read_auxiliar_sft(path: Path | str, aba: str = "SFT ") -> dict[str, FornecedorLookup]:
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = aba if aba in wb.sheetnames else aba.strip()
    if sheet not in wb.sheetnames:
        for name in wb.sheetnames:
            if name.strip().upper() == aba.strip().upper():
                sheet = name
                break
        else:
            wb.close()
            return {}

    ws = wb[sheet]
    header_row, hdr = _find_sft_header(ws)
    idx, idx_norm = _build_header_index(hdr)
    chave_col = _chave_col_index(idx, idx_norm)
    nome_col = _nome_forn_col_index(idx, idx_norm)
    dt_emissao_col = _date_col_index(idx, idx_norm, emissao=True)
    dt_entrada_col = _date_col_index(idx, idx_norm, emissao=False)
    idx_nota = _sft_component_col(idx, idx_norm, ("Nota Fiscal", "NOTA FISCAL", "Nota"))
    idx_forn = _sft_component_col(
        idx, idx_norm, ("Cd. Forn", "Cd Forn", "Cod Fornecedor", "COD FORNECEDOR")
    )
    idx_prod = _sft_component_col(
        idx, idx_norm, ("Cd Produto", "Cd. Produto", "Cód Produto", "COD PRODUTO")
    )
    lookups: dict[str, FornecedorLookup] = {}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        cells = list(row)
        nota = cells[idx_nota] if idx_nota is not None and idx_nota < len(cells) else None
        forn = normalize_codigo(_cell_str(cells, idx_forn))
        prod = normalize_codigo(cells[idx_prod]) if idx_prod is not None and idx_prod < len(cells) else ""
        chave_kgq = build_chave_sft_row(cells)
        chave_header = (
            build_chave_rastreio(str(nota), forn, prod) if nota not in (None, "") and forn and prod else ""
        )
        chave_sft = chave_from_sft_cell(cells[chave_col] if chave_col < len(cells) else None)
        chave_primary = chave_kgq or chave_header or chave_sft
        if not chave_primary:
            continue

        data_emissao = (
            _parse_date(cells[dt_emissao_col])
            if dt_emissao_col is not None and dt_emissao_col < len(cells)
            else None
        )
        data_entrada = (
            _parse_date(cells[dt_entrada_col])
            if dt_entrada_col is not None and dt_entrada_col < len(cells)
            else None
        )

        nome_forn = _cell_str(cells, nome_col)
        if not nome_forn and nome_col != SFT_NOME_COL:
            nome_forn = _cell_str(cells, SFT_NOME_COL)

        lookup = FornecedorLookup(
            chave_rastreio=chave_primary,
            nome_fornecedor=nome_forn,
            loja=(_cell_str(cells, idx.get("Lj. Forn")) or "0001").zfill(4),
            data_emissao=data_emissao,
            data_entrada=data_entrada,
            cod_item_contabil=_cell_str(cells, idx.get("COD ITEM CONTABIL")),
        )

        for key in {chave_kgq, chave_header, chave_sft, normalize_chave(chave_kgq), normalize_chave(chave_sft)}:
            if key:
                lookups[key] = lookup
    wb.close()
    return lookups
