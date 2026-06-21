from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from difal_importacao.branding import logo_path
from difal_importacao.models import LancamentoImportacao, TotaisConsolidados

IMPORT_HEADER_ROW = 4
IMPORT_DATA_START_ROW = 5
IMPORT_NUM_COLS = 27

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
STRIPE_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN_SIDE = Side(style="thin", color="8EA9DB")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def format_data_br(val: datetime | date | None) -> str | None:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    return str(val)


IMPORT_HEADERS = [
    "LOJA", "NOTA", "NOME FORNE", "COD FORNECEDOR", "PRODUTO", "CFOP",
    "VALOR DA NOTA", "DIFAL", "NOVO DIFAL", "AJUSTE", "Chave",
    "DATA EMISSAO", "DATA ENTRADA", "FILIAL", "DÉBITO", "CENTRO DE CUSTO",
    "COD ITEM CONTABIL", "COD FORNECEDOR", "CRÉDITO", "CENTRO DE CUSTO",
    "COD ITEM CONTABIL", "COD FORNECEDOR", "VALOR", "HISTÓRICO",
    "JUSTIFICATIVA", "DATA EMISSÃO", "DATA CONTABILIZAÇÃO",
]


def _insert_logo(ws) -> None:
    path = logo_path()
    if not path:
        return
    img = Image(str(path))
    max_w = 200
    if img.width and img.width > max_w:
        ratio = max_w / img.width
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)
    ws.add_image(img, "A1")
    ws.row_dimensions[1].height = max(52, (img.height or 52) * 0.75)


def _format_importacao_table(ws, header_row: int, last_row: int) -> None:
    if last_row < header_row:
        return
    last_col = IMPORT_NUM_COLS
    last_letter = get_column_letter(last_col)

    for col in range(1, last_col + 1):
        cell = ws.cell(header_row, col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(header_row + 1, last_row + 1):
        stripe = (row_idx - header_row) % 2 == 0
        for col in range(1, last_col + 1):
            cell = ws.cell(row_idx, col)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if stripe:
                cell.fill = STRIPE_FILL

    ws.auto_filter.ref = f"A{header_row}:{last_letter}{last_row}"
    ws.freeze_panes = f"A{IMPORT_DATA_START_ROW}"


def write_importacao_sheet(
    lancamentos: list[LancamentoImportacao],
    totais: TotaisConsolidados,
    output: Path | str,
    source_workbook: Path | str | None = None,
) -> Path:
    output = Path(output)
    if source_workbook and Path(source_workbook).exists():
        wb = openpyxl.load_workbook(source_workbook)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "INDUSTRIA-IMPORTAÇÃO" in wb.sheetnames:
        del wb["INDUSTRIA-IMPORTAÇÃO"]
    ws = wb.create_sheet("INDUSTRIA-IMPORTAÇÃO")

    ws.append([None] * IMPORT_NUM_COLS)
    row2 = [None] * IMPORT_NUM_COLS
    row2[6] = totais.total_valor_nota
    row2[7] = totais.total_difal
    row2[8] = totais.total_novo_difal
    row2[9] = totais.total_ajuste
    ws.append(row2)
    ws.append([None] * IMPORT_NUM_COLS)
    ws.append(IMPORT_HEADERS)

    for l in lancamentos:
        emissao = format_data_br(l.data_emissao)
        entrada = format_data_br(l.data_entrada)
        ws.append([
            l.loja, l.nota, l.nome_fornecedor, l.cod_fornecedor, l.produto, l.cfop,
            l.valor_nota, l.difal, l.novo_difal, l.ajuste, l.chave,
            emissao, entrada, l.filial,
            l.debito, l.centro_custo, l.cod_item_contabil, l.cod_fornecedor_debito,
            l.credito, l.centro_custo, l.cod_item_contabil, l.cod_fornecedor_credito,
            l.valor_lancamento, l.historico, l.justificativa,
            emissao, entrada,
        ])

    last_row = ws.max_row
    _insert_logo(ws)
    _format_importacao_table(ws, IMPORT_HEADER_ROW, last_row)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    wb.close()
    return output
