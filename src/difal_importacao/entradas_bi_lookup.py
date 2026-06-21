"""Lookup COD ITEM CONTABIL por código de produto — aba ENTRADAS-BI, coluna AK."""
from functools import lru_cache
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

from difal_apuracao.sb1_lookup import _find_sb1_workbook, normalize_codigo

ENTRADAS_BI_SHEET = "ENTRADAS-BI"
COD_ITEM_COL = column_index_from_string("AK") - 1  # 0-based
PRODUCT_HEADERS = ("Cód Produto", "Cod Produto", "COD PRODUTO", "Cd Produto")


def _find_product_col(hdr_row: tuple) -> int:
    idx = {str(h).strip(): i for i, h in enumerate(hdr_row) if h}
    for name in PRODUCT_HEADERS:
        if name in idx:
            return idx[name]
    return 10


def _format_item(val) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, (int, float)):
        if val == int(val):
            return str(int(val))
        return str(val).strip()
    return str(val).strip()


@lru_cache(maxsize=2)
def load_entradas_item_contabil(workbook: str) -> dict[str, str]:
    path = Path(workbook)
    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if ENTRADAS_BI_SHEET not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb[ENTRADAS_BI_SHEET]
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    prod_col = _find_product_col(hdr)
    mapping: dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        cells = list(row)
        if prod_col >= len(cells):
            continue
        if cells[prod_col] is None:
            continue
        produto = normalize_codigo(cells[prod_col])
        if not produto:
            continue
        item_val = cells[COD_ITEM_COL] if COD_ITEM_COL < len(cells) else None
        item = _format_item(item_val)
        if item:
            mapping[produto] = item

    wb.close()
    return mapping


def load_entradas_map(
    workbook: Path | str | None = None,
    fallback_workbook: Path | str | None = None,
) -> dict[str, str]:
    for candidate in (workbook, fallback_workbook):
        if not candidate:
            continue
        path = Path(candidate)
        if path.exists():
            result = load_entradas_item_contabil(str(path))
            if result:
                return result

    auto = _find_sb1_workbook(Path(workbook) if workbook else None)
    if auto:
        return load_entradas_item_contabil(str(auto))
    return {}
