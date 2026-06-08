"""Escrita da aba DIFAL MM.AAAA."""
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from difal_apuracao.models import LinhaDifal

DIFAL_HEADERS = [
    "FORNECEDOR", "ESTADO", "COD FILIAL", "NOTA FISCAL", "CONTA CONTABIL",
    "COD PRODUTO", "PRODUTO", "NCM", "COD FISCAL", "QUANTIDADE", "PREÇO UNITÁRIO",
    "TOTAL", "DESCONTO", "DESPESAS", "FRETE", "VALOR IPI ", "ICMS RETIDO",
    "VALOR CONTÁBIL", "ALÍQUOTA ICMS", "VALOR ICMS", "ALÍQUOTA COMPLEMENTAR",
    "ALIQUOTA ICMS COMPLEMENTAR", "VALOR ICMS COMPLEMENTAR", "NOVO DIFAL", "AJUSTE",
    "METODO CALCULO", "NCM REGRA", "CARGA NORMATIVA NCM", "MEMORIA CALCULO",
]


def write_difal_sheet(
    linhas: list[LinhaDifal],
    output: Path | str,
    periodo_label: str,
    source_workbook: Path | str | None = None,
) -> Path:
    output = Path(output)
    if source_workbook and Path(source_workbook).exists():
        wb = openpyxl.load_workbook(source_workbook)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_name = f"DIFAL {periodo_label} "
    if sheet_name.strip() in [s.strip() for s in wb.sheetnames]:
        for name in wb.sheetnames:
            if name.strip() == sheet_name.strip():
                del wb[name]
                break
    ws = wb.create_sheet(sheet_name)
    ws.append(DIFAL_HEADERS)

    for l in linhas:
        ws.append([
            l.fornecedor, l.estado, l.cod_filial, l.nota_fiscal, l.conta_contabil,
            l.cod_produto, l.produto, l.ncm, l.cod_fiscal, l.quantidade,
            l.preco_unitario, l.total, l.desconto, l.despesas, l.frete,
            l.valor_ipi, l.icms_retido, l.valor_contabil, l.aliquota_icms,
            l.valor_icms, l.aliquota_complementar, l.aliquota_icms_complementar,
            l.valor_icms_complementar, l.novo_difal, l.ajuste,
            l.metodo_calculo, l.ncm_regra_aplicada, l.carga_normativa_ncm, l.memoria_calculo,
        ])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    wb.close()
    return output
