"""Lookup conta contábil (Cta Custo) por código de produto — planilha SB1."""
from functools import lru_cache
from pathlib import Path

import openpyxl

SB1_SHEET = "SB1 - PRODUTO X CONTA CONTÁBIL"
HEADER_ROW = 3
DATA_START = 4


def normalize_codigo(val) -> str:
    s = str(val).strip()
    if s.endswith(".0"):
        head = s[:-2]
        if head:
            return head
    return s


def normalize_cfop(val) -> str:
    s = str(val).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    try:
        num = float(s.replace(",", "."))
        if num == int(num):
            return str(int(num))
    except ValueError:
        pass
    return s


def _find_sb1_workbook(explicit: Path | None) -> Path | None:
    if explicit and explicit.exists():
        return explicit
    import os

    roots = []
    if os.environ.get("DIFAL_ROOT"):
        roots.append(Path(os.environ["DIFAL_ROOT"]))
    roots.append(Path(__file__).resolve().parents[2])
    for root in roots:
        candidates = list(root.glob("*28*.xlsx"))
        if candidates:
            return candidates[0]
    return None


@lru_cache(maxsize=2)
def load_sb1_contas(workbook: str) -> dict[str, str]:
    path = Path(workbook)
    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SB1_SHEET not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb[SB1_SHEET]
    mapping: dict[str, str] = {}
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        if not row or row[0] is None:
            continue
        codigo = normalize_codigo(row[0])
        cta_custo = row[3]
        if cta_custo is not None and str(cta_custo).strip():
            conta = str(int(cta_custo)) if isinstance(cta_custo, (int, float)) else str(cta_custo).strip()
            mapping[codigo] = conta
    wb.close()
    return mapping


def resolve_conta_produto(
    cod_produto: str,
    sb1_workbook: Path | str | None = None,
    conta_por_grupo: dict[str, str] | None = None,
    grupo: str = "",
    conta_por_cfop: dict[str, str | None] | None = None,
    cfop: str = "",
) -> str:
    wb_path = _find_sb1_workbook(Path(sb1_workbook) if sb1_workbook else None)
    if wb_path:
        conta = load_sb1_contas(str(wb_path)).get(normalize_codigo(cod_produto))
        if conta:
            return conta

    if grupo and conta_por_grupo:
        conta = conta_por_grupo.get(grupo.strip())
        if conta:
            return conta

    if conta_por_cfop:
        conta = conta_por_cfop.get(str(cfop).strip())
        if conta:
            return conta
        if str(cfop).strip() == "2551":
            return "12050010100"

    return "#N/A"
