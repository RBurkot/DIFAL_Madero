"""Leitura do extrato BI (DIFAL INDUSTRIA BI.xlsx)."""
from datetime import datetime
from pathlib import Path

import openpyxl

from difal_apuracao.config import ApuracaoConfig, load_config
from difal_apuracao.models import LinhaBI

BI_REQUIRED = [
    "MES_ANO_ENTRADA",
    "Cod Fornecedor",
    "Estado",
    "Cód Filial",
    "Nota Fiscal",
    "Cód Produto",
    "Produto",
    "NCM",
    "Cód Fiscal",
    "Valor Contábil",
    "Alíquota ICMS",
    "Alíquota Compl.",
    "Aliquota ICMS Complementar",
    "D1_ICMSCOM",
]


class BIValidationError(Exception):
    pass


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row and row[1] == "Cod Fornecedor":
            idx = {str(h): i for i, h in enumerate(row) if h}
            missing = [c for c in BI_REQUIRED if c not in idx]
            if missing:
                raise BIValidationError(f"Colunas obrigatórias ausentes: {', '.join(missing)}")
            return row_idx, idx
    raise BIValidationError("Cabeçalho BI não encontrado (esperado linha com 'Cod Fornecedor')")


def _num(val, default=0.0) -> float:
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def read_bi(
    path: Path | str,
    periodo_mes: int | None = None,
    periodo_ano: int | None = None,
    corte_dia: int | None = None,
    config: ApuracaoConfig | None = None,
) -> list[LinhaBI]:
    cfg = config or load_config()
    cfops = {str(c).strip() for c in cfg.cfops_escopo} if cfg.cfops_escopo else None
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, idx = _find_header_row(ws)
    linhas: list[LinhaBI] = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[idx["Cod Fornecedor"]]:
            continue
        mes_entrada = row[idx["MES_ANO_ENTRADA"]]
        if periodo_mes and periodo_ano and isinstance(mes_entrada, datetime):
            if mes_entrada.month != periodo_mes or mes_entrada.year != periodo_ano:
                continue
            if corte_dia and mes_entrada.day > corte_dia:
                continue

        cfop = str(row[idx["Cód Fiscal"]]).strip()
        if cfops and cfop not in cfops:
            continue

        carga = None
        if "Carga Efetiva DIFAL" in idx:
            carga = _num(row[idx["Carga Efetiva DIFAL"]], default=None)  # type: ignore

        linhas.append(
            LinhaBI(
                mes_ano_entrada=mes_entrada if isinstance(mes_entrada, datetime) else None,
                cod_fornecedor=str(row[idx["Cod Fornecedor"]]).strip(),
                estado=str(row[idx["Estado"]]).strip(),
                cod_filial=str(row[idx["Cód Filial"]]).strip(),
                nota_fiscal=str(row[idx["Nota Fiscal"]]).strip(),
                cod_produto=str(row[idx["Cód Produto"]]).strip(),
                produto=str(row[idx.get("Produto", idx["Cód Produto"])] or "").strip(),
                ncm=str(row[idx["NCM"]]).strip(),
                cod_fiscal=str(row[idx["Cód Fiscal"]]).strip(),
                quantidade=_num(row[idx.get("Quantidade", 0)] if "Quantidade" in idx else 0),
                preco_unitario=_num(row[idx.get("Preço Unitário", 0)] if "Preço Unitário" in idx else 0),
                total=_num(row[idx.get("Total", 0)] if "Total" in idx else 0),
                valor_contabil=_num(row[idx["Valor Contábil"]]),
                aliquota_icms=_num(row[idx["Alíquota ICMS"]]),
                valor_icms=_num(row[idx.get("Valor ICMS", 0)] if "Valor ICMS" in idx else 0),
                aliquota_compl=_num(row[idx["Alíquota Compl."]]),
                aliquota_icms_complementar=_num(row[idx["Aliquota ICMS Complementar"]]),
                valor_icms_complementar=_num(
                    row[idx.get("Valor ICMS Complementar", 0)] if "Valor ICMS Complementar" in idx else 0
                ),
                carga_efetiva_difal=carga,
                d1_icmscom=_num(row[idx["D1_ICMSCOM"]]),
                desc_grupo=str(row[idx.get("Desc. Grupo Produtos", "")] or "") if "Desc. Grupo Produtos" in idx else "",
                desc_tes=str(row[idx.get("Descr. TES", "")] or "") if "Descr. TES" in idx else "",
            )
        )
    wb.close()
    return linhas


def validate_bi_layout(path: Path | str) -> tuple[bool, list[str]]:
    try:
        path = Path(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        _find_header_row(ws)
        wb.close()
        return True, []
    except BIValidationError as e:
        return False, [str(e)]
    except Exception as e:
        return False, [f"Erro ao ler arquivo: {e}"]
